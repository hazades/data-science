from urllib.request import urlopen
from bs4 import BeautifulSoup
import os
import time
import requests
import re
import openpyxl
from pathlib import Path

url = "https://parking.klcc.com.my/"
filename = "parking_klcc.xlsx"
pt = Path(filename)
cp = 1
ct = 1

def timestamp():
	timestamp = ''
	for i in content.findAll('h3'):
		timestamp = timestamp + ' ' + i.text
		ts = timestamp.split(" ")[4]+" "+ timestamp.split(" ")[5]
		# adding timestamp value inside excel
		c1 = sheet.cell(row = ct, column = 1)
		c1.value = ts
	print(ts)

def parking_name(cp):
	parking =  []
	for i in content.findAll('h4'):
		p = i.text
		#print(p)
		parking.append(p)
	#print(parking)
	for j in range(len(parking)):
		print(parking[j])
		c2 = sheet.cell(row = cp, column = 4)
		c2.value = parking[j]
		cp+=1

def total(ct):
	#read from linux command also can
	#value = os.popen("curl -s https://parking.klcc.com.my/ | grep -w 'value' | grep -v 'text'| sed 's/[^0-9]*//g'").read()
	value = requests.get("https://parking.klcc.com.my/")
	value_lines = value.text.splitlines()
	value_array = []

	for line in value_lines:
		if re.findall("value", line) and not re.findall("text", line):
			if not re.findall("valueStrokeClass", line):
				value_array.append(re.findall("([0-9]+)", line)[0])
	#print(value_array)
	
	for tt in range(len(value_array)):
		print(value_array[tt])
	
	for i in range(1, len(value_array)+1):
		c3 = sheet.cell(row = ct, column = 8)
		c3.value = value_array[i-1]
		ct+=1

try:
	while True:
		print("\nThis App will keep refreshing in 15 seconds:")
		print("Kindly press ctrl + z to terminate the program\n")

		page = urlopen(url)
		soup = BeautifulSoup(page, 'html.parser')
		content = soup.find('div', {"class": "contentgray"})
		
		if pt.is_file():
			print ("File exist")
			wb = openpyxl.load_workbook(filename)
			sheet = wb.active
		else:
			print ("File not exist")
			wb = openpyxl.Workbook()
			sheet = wb.active
			sheet.title = "Parking Midvalley"
		
		#calling functions
		timestamp()
		parking_name(cp)
		total(ct)
		
		#algo to increase the row number for excel
		cp+=12
		ct+=12
		
		#save file inside excel
		wb.save(filename=filename)
		
		#loop every 15 seconds
		time.sleep(30)

except KeyboardInterrupt:
	print("System Exitting")
