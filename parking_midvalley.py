from urllib.request import urlopen
from bs4 import BeautifulSoup
import time
import openpyxl
from pathlib import Path

url = "http://midvalley.com.my/locate/by-car/"
filename = "parking_midvalley.xlsx"
pt = Path(filename)
r = 1

def total(r):
	data = []
	for i in content.find_all('td'):
		d = i.get_text()
		#insert all data inside array named data[]
		data.append(d)
	for j in range(len(data)):
		print(data[j])
		c1 = sheet.cell(row = r, column = 1)
		c1.value = data[j]
		r+=1

try:
	while True:
		print("\nThis App will keep refreshing in 30 seconds:")
		print("Kindly press ctrl + z or exit to terminate the program\n")

		page = urlopen(url)
		soup = BeautifulSoup(page, 'html.parser')
		content = soup.find('div', {"class": "col-sm-3"})
		
		if pt.is_file():
			print ("File exist")
			wb = openpyxl.load_workbook(filename)
			sheet = wb.active
		else:
			print ("File not exist")
			wb = openpyxl.Workbook()
			sheet = wb.active
			sheet.title = "Parking Midvalley"
		
		#calling functions
		total(r)
		#algo to increase the row number for excel
		r+=17
		#save file inside excel
		wb.save(filename=filename)
		#loop every 15 seconds
		time.sleep(30)

except KeyboardInterrupt:
	print("System Exitting")
